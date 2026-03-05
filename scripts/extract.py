"""
Extrai KPIs das planilhas de séries históricas dos bancos e gera JSON unificado.
Bancos: Itaú, Bradesco, BB, BTG, Santander
"""

import openpyxl
import json
import re
from pathlib import Path

WORKSPACE = Path(__file__).parent.parent.parent  # workspace-bruno
DATA_OUT = Path(__file__).parent.parent / "data"
DATA_OUT.mkdir(exist_ok=True)


def safe_float(val):
    if val is None or val == "" or val == "-" or val == "n/a":
        return None
    try:
        f = float(val)
        return round(f, 6)
    except (ValueError, TypeError):
        return None


def normalize_period(raw):
    """Converte períodos para formato padrão YYYY-QN (ex: '1T25' -> '2025-Q1')"""
    if raw is None:
        return None
    s = str(raw).strip()
    # Formato nT YY (ex: 1T25, 4T19)
    m = re.match(r'^(\d)T(\d{2})$', s)
    if m:
        q, yr = m.group(1), m.group(2)
        year = 2000 + int(yr)
        return f"{year}-Q{q}"
    # Formato NQ YYYY (ex: 1Q 2025, 4Q 2011)
    m = re.match(r'^(\d)Q\s*(\d{4})$', s)
    if m:
        q, year = m.group(1), m.group(2)
        return f"{year}-Q{q}"
    # Formato Mmm/YY (ex: Mar25, Dez07)
    months = {'Jan':1,'Fev':2,'Mar':3,'Abr':4,'Mai':5,'Jun':6,
              'Jul':7,'Ago':8,'Set':9,'Out':10,'Nov':11,'Dez':12,
              'Feb':2,'Apr':4,'May':5,'Aug':8,'Sep':9,'Oct':10,'Dec':12}
    m = re.match(r'^([A-Za-z]{3})[/\-]?(\d{2,4})$', s)
    if m:
        mon_str, yr_str = m.group(1).capitalize(), m.group(2)
        yr = int(yr_str) if len(yr_str) == 4 else 2000 + int(yr_str)
        mon = months.get(mon_str, 0)
        if mon:
            q = (mon - 1) // 3 + 1
            return f"{yr}-Q{q}"
    # Formato Mmm/YYYY
    m = re.match(r'^([A-Za-z]{3})/(\d{4})$', s)
    if m:
        mon_str, yr = m.group(1).capitalize(), int(m.group(2))
        mon = months.get(mon_str, 0)
        if mon:
            q = (mon - 1) // 3 + 1
            return f"{yr}-Q{q}"
    return None


# ─────────────────────────────────────────────
# ITAÚ
# ─────────────────────────────────────────────
def extract_itau():
    path = WORKSPACE / "itau_series_historicas_4T25.xlsx"
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    ws = wb["Sumário_PRO FORMA"]
    rows = list(ws.iter_rows(values_only=True))

    # Linha 2 tem os períodos (colunas 3+)
    header_row = rows[1]
    periods = []
    period_cols = []
    for i, val in enumerate(header_row):
        p = normalize_period(val)
        if p:
            periods.append(p)
            period_cols.append(i)

    def get_series(label_fragment, rows, scale=1.0):
        for row in rows:
            if row[1] and label_fragment.lower() in str(row[1]).lower():
                return {periods[j]: safe_float(row[period_cols[j]]) and round(safe_float(row[period_cols[j]]) * scale, 4)
                        for j in range(len(periods))}
        return {}

    data = {
        "banco": "Itaú Unibanco",
        "ticker": "ITUB4",
        "kpis": {
            "lucro_liquido_recorrente": get_series("Resultado Recorrente Gerencial", rows, 1/1000),  # -> R$ bi
            "roe": get_series("Retorno Recorrente Gerencial s", rows),
            "indice_eficiencia": get_series("Índice de Eficiência", rows),
            "carteira_credito": get_series("Total de Operações de Crédito", rows, 1/1000),
            "total_ativos": get_series("Ativos Totais", rows, 1/1000),
            "pl": get_series("Patrimônio Líquido", rows, 1/1000),
            "inadimplencia_90d": get_series("Índice de Inadimplência (90 di", rows),
            "basileia": get_series("Índice de Basileia", rows),
        }
    }

    # Margem Financeira Gerencial
    mf = get_series("Margem Financeira Gerencial", rows, 1/1000)
    data["kpis"]["margem_financeira"] = mf

    # NIM sheet para receita de serviços
    ws2 = wb["Sumário_PRO FORMA"]
    rows2 = list(ws2.iter_rows(values_only=True))
    data["kpis"]["receita_servicos"] = get_series("Receitas de Prestação de Serviços", rows2, 1/1000)

    wb.close()
    return data


# ─────────────────────────────────────────────
# BRADESCO
# ─────────────────────────────────────────────
def extract_bradesco():
    path = WORKSPACE / "bradesco_series_historicas_4T25.xlsx"
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    # KPIs de desempenho (longo histórico)
    ws_desempenho = wb["27- Índice de Desempenho"]
    rows_d = list(ws_desempenho.iter_rows(values_only=True))
    header = rows_d[6]
    periods = []
    period_cols = []
    for i, val in enumerate(header):
        p = normalize_period(val)
        if p:
            periods.append(p)
            period_cols.append(i)

    def get_series_d(label_fragment):
        for row in rows_d[7:]:
            if row[0] and label_fragment.lower() in str(row[0]).lower():
                result = {}
                for j, col in enumerate(period_cols):
                    v = safe_float(row[col])
                    if v is not None:
                        result[periods[j]] = v
                return result
        return {}

    # DRE recorrente (apenas 2025)
    ws_dre = wb["4- DRE Recorrente"]
    rows_dre = list(ws_dre.iter_rows(values_only=True))
    header_dre = rows_dre[6]
    periods_dre = []
    cols_dre = []
    for i, val in enumerate(header_dre):
        p = normalize_period(val)
        if p:
            periods_dre.append(p)
            cols_dre.append(i)

    def get_series_dre(label_fragment):
        for row in rows_dre[8:]:
            if row[0] and label_fragment.lower() in str(row[0]).lower():
                return {periods_dre[j]: safe_float(row[cols_dre[j]]) and round(safe_float(row[cols_dre[j]]) / 1000, 4)
                        for j in range(len(periods_dre))}
        return {}

    # BP dados selecionados (apenas 2025)
    ws_bp = wb["1- BP - Dados Selecionados "]
    rows_bp = list(ws_bp.iter_rows(values_only=True))
    header_bp = rows_bp[6]
    periods_bp = []
    cols_bp = []
    for i, val in enumerate(header_bp):
        p = normalize_period(val)
        if p:
            periods_bp.append(p)
            cols_bp.append(i)

    def get_series_bp(label_fragment):
        for row in rows_bp[8:]:
            if row[0] and label_fragment.lower() in str(row[0]).lower():
                return {periods_bp[j]: safe_float(row[cols_bp[j]]) and round(safe_float(row[cols_bp[j]]) / 1000, 4)
                        for j in range(len(periods_bp))}
        return {}

    # ROE histórico (em %) - converter de decimal se necessário
    roe_raw = get_series_d("Retorno Anualizado sob")
    roe = {}
    for k, v in roe_raw.items():
        if v is not None:
            roe[k] = round(v / 100, 6) if v > 1 else v

    ie_raw = get_series_d("Índice de Eficiência")
    ie = {}
    for k, v in ie_raw.items():
        if v is not None:
            ie[k] = round(v / 100, 6) if v > 1 else v

    data = {
        "banco": "Bradesco",
        "ticker": "BBDC4",
        "kpis": {
            "roe": roe,
            "indice_eficiencia": ie,
            "lucro_liquido_recorrente": get_series_dre("Lucro Líquido Recorrente"),
            "margem_financeira": get_series_dre("Margem  Financeira"),
            "receita_servicos": get_series_dre("Receitas de Prest"),
            "total_ativos": get_series_bp("Total de ativos"),
            "pl": get_series_bp("Patrimônio"),
            "carteira_credito": get_series_bp("Operações de crédito"),
        }
    }

    # Inadimplência - aba Índice Cobertura
    ws_cob = wb["29- Índice Cobertura"]
    rows_cob = list(ws_cob.iter_rows(values_only=True))
    wb.close()
    return data


# ─────────────────────────────────────────────
# BANCO DO BRASIL
# ─────────────────────────────────────────────
def extract_bb():
    path = WORKSPACE / "bb_series_historicas_4T25.xlsx"
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    # Desempenho e Eficiência
    ws_desemp = wb["Desempenho e Eficiência"]
    rows_d = list(ws_desemp.iter_rows(values_only=True))
    header = rows_d[3]
    periods = []
    period_cols = []
    for i, val in enumerate(header):
        p = normalize_period(val)
        if p:
            periods.append(p)
            period_cols.append(i)

    def get_series_d(label_fragment, scale=1.0):
        for row in rows_d[5:]:
            if row[0] and label_fragment.lower() in str(row[0]).lower():
                result = {}
                for j, col in enumerate(period_cols):
                    v = safe_float(row[col])
                    if v is not None:
                        result[periods[j]] = round(v * scale, 6)
                return result
        return {}

    # DRE com Realocações (apenas 2025)
    ws_dre = wb["DRE com Realocações"]
    rows_dre = list(ws_dre.iter_rows(values_only=True))
    header_dre = rows_dre[3]
    periods_dre = []
    cols_dre = []
    for i, val in enumerate(header_dre):
        p = normalize_period(val)
        if p:
            periods_dre.append(p)
            cols_dre.append(i)

    def get_series_dre(label_fragment, scale=1.0):
        for row in rows_dre[5:]:
            if row[0] and label_fragment.lower() in str(row[0]).lower():
                return {periods_dre[j]: safe_float(row[cols_dre[j]]) and round(safe_float(row[cols_dre[j]]) * scale, 4)
                        for j in range(len(periods_dre))}
        return {}

    # Carteira de Crédito
    ws_cart = wb["Carteira de Crédito"]
    rows_cart = list(ws_cart.iter_rows(values_only=True))
    header_cart = rows_cart[3]
    periods_cart = []
    cols_cart = []
    for i, val in enumerate(header_cart):
        p = normalize_period(val)
        if p:
            periods_cart.append(p)
            cols_cart.append(i)

    def get_series_cart(label_fragment):
        for row in rows_cart[5:]:
            if row[0] and label_fragment.lower() in str(row[0]).lower():
                return {periods_cart[j]: safe_float(row[cols_cart[j]]) and round(safe_float(row[cols_cart[j]]) / 1e9, 4)
                        for j in range(len(periods_cart))}
        return {}

    # BP Ativo
    ws_bp = wb["BP - Ativo"]
    rows_bp = list(ws_bp.iter_rows(values_only=True))
    header_bp = rows_bp[3] if len(rows_bp) > 3 else []
    periods_bp = []
    cols_bp = []
    for i, val in enumerate(header_bp):
        p = normalize_period(val)
        if p:
            periods_bp.append(p)
            cols_bp.append(i)

    def get_series_bp(label_fragment):
        for row in rows_bp[4:]:
            if row[0] and label_fragment.lower() in str(row[0]).lower():
                return {periods_bp[j]: safe_float(row[cols_bp[j]]) and round(safe_float(row[cols_bp[j]]) / 1e9, 4)
                        for j in range(len(periods_bp))}
        return {}

    # ROE e IE podem estar em % já
    roe_raw = get_series_d("Rentabilidade s/ PL")
    roe = {k: round(v / 100, 6) if v is not None and v > 1 else v for k, v in roe_raw.items()}

    ie_raw = get_series_d("Índice de Eficiência")
    ie = {k: round(v / 100, 6) if v is not None and v > 1 else v for k, v in ie_raw.items()}

    # Lucro em R$ bilhões (DRE está em unidades)
    lucro_dre = get_series_dre("Lucro Líquido", scale=1/1e9)

    # Margem financeira bruta (em unidades -> bilhões)
    mf_dre = get_series_dre("Margem Financeira Brut", scale=1/1e9)

    data = {
        "banco": "Banco do Brasil",
        "ticker": "BBAS3",
        "kpis": {
            "roe": roe,
            "indice_eficiencia": ie,
            "lucro_liquido_recorrente": lucro_dre,
            "margem_financeira": mf_dre,
            "carteira_credito": get_series_cart("Carteira de Crédito"),
            "total_ativos": get_series_bp("Total do Ativo"),
        }
    }

    # Basileia
    ws_bas = wb["Índice de Basileia"]
    rows_bas = list(ws_bas.iter_rows(values_only=True))
    header_bas = rows_bas[3] if len(rows_bas) > 3 else []
    periods_bas = []
    cols_bas = []
    for i, val in enumerate(header_bas):
        p = normalize_period(val)
        if p:
            periods_bas.append(p)
            cols_bas.append(i)
    for row in rows_bas[4:]:
        if row[0] and "basileia" in str(row[0]).lower():
            bas = {}
            for j, col in enumerate(cols_bas):
                v = safe_float(row[col])
                if v is not None:
                    bas[periods_bas[j]] = round(v / 100, 6) if v > 1 else v
            data["kpis"]["basileia"] = bas
            break

    wb.close()
    return data


# ─────────────────────────────────────────────
# BTG PACTUAL
# ─────────────────────────────────────────────
def extract_btg():
    path = WORKSPACE / "btg_series_historicas_4T25.xlsx"
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    ws = wb["InputSite_Highlights"]
    rows = list(ws.iter_rows(values_only=True))

    header = rows[1]
    periods = []
    period_cols = []
    for i, val in enumerate(header):
        p = normalize_period(val)
        if p:
            periods.append(p)
            period_cols.append(i)

    def get_series(label_fragment, scale=1.0):
        for row in rows:
            if row[0] and label_fragment.lower() in str(row[0]).lower():
                result = {}
                for j, col in enumerate(period_cols):
                    v = safe_float(row[col])
                    if v is not None:
                        result[periods[j]] = round(v * scale, 6)
                return result
        return {}

    # Income Statement
    ws2 = wb["InputSite_IncomeStatement"]
    rows2 = list(ws2.iter_rows(values_only=True))

    header2 = rows2[1]
    periods2 = []
    cols2 = []
    for i, val in enumerate(header2):
        p = normalize_period(val)
        if p:
            periods2.append(p)
            cols2.append(i)

    def get_series2(label_fragment, scale=1.0):
        for row in rows2:
            if row[0] and label_fragment.lower() in str(row[0]).lower():
                result = {}
                for j, col in enumerate(cols2):
                    v = safe_float(row[col])
                    if v is not None:
                        result[periods2[j]] = round(v * scale, 6)
                return result
        return {}

    # ROE e IE já em decimal (< 1)
    roe_raw = get_series("ROAE Anualizado")
    ie_raw = get_series("Índice de eficiência")

    data = {
        "banco": "BTG Pactual",
        "ticker": "BPAC11",
        "kpis": {
            "lucro_liquido_recorrente": get_series("Lucro Liquido Ajustado", scale=1/1000),
            "roe": roe_raw,
            "indice_eficiencia": ie_raw,
            "total_ativos": get_series("Total Ativos", scale=1/1000),
            "pl": get_series("Patrimônio Liquido", scale=1/1000),
            "basileia": get_series("Índice de Basiléia"),
            "aum": get_series("AuM e AuA"),
        }
    }

    wb.close()
    return data


# ─────────────────────────────────────────────
# SANTANDER
# ─────────────────────────────────────────────
def extract_santander():
    path = WORKSPACE / "santander_series_historicas_4T25.xlsx"
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    def extract_from_sheet(sheet_name, label_fragment, scale=1.0):
        try:
            ws = wb[sheet_name]
        except KeyError:
            return {}
        rows = list(ws.iter_rows(values_only=True))
        # Encontrar linha de períodos
        period_row_idx = None
        for i, row in enumerate(rows[:15]):
            trimestres = [normalize_period(c) for c in row if normalize_period(c)]
            if len(trimestres) >= 2:
                period_row_idx = i
                break
        if period_row_idx is None:
            return {}
        header = rows[period_row_idx]
        periods = []
        period_cols = []
        for i, val in enumerate(header):
            p = normalize_period(val)
            if p:
                periods.append(p)
                period_cols.append(i)
        for row in rows[period_row_idx+1:]:
            if row[1] and label_fragment.lower() in str(row[1]).lower():
                result = {}
                for j, col in enumerate(period_cols):
                    v = safe_float(row[col])
                    if v is not None:
                        result[periods[j]] = round(v * scale, 6)
                return result
        return {}

    # Consolidar séries históricas gerenciais (múltiplas abas)
    def merge_series(*dicts):
        merged = {}
        for d in dicts:
            for k, v in d.items():
                if k not in merged and v is not None:
                    merged[k] = v
        return dict(sorted(merged.items()))

    # ROE e IE nos indicadores históricos
    roe_2022 = extract_from_sheet("Indicadores Ger. até 2022", "ROE")
    roe_20242 = extract_from_sheet("Indicadores Ger. 2023 a 2024", "ROE")
    roe_curr = extract_from_sheet("Sumário Executivo", "ROE")

    ie_2022 = extract_from_sheet("Indicadores Ger. até 2022", "Eficiência")
    ie_2024 = extract_from_sheet("Indicadores Ger. 2023 a 2024", "Eficiência")
    ie_curr = extract_from_sheet("Sumário Executivo", "Eficiência")

    # DRE Gerencial histórica
    ll_old = extract_from_sheet("DRE_Gerencial até 2018", "Lucro líquido", scale=1/1e6)
    ll_1922 = extract_from_sheet("DRE_Gerencial 2019 a 2022", "Lucro líquido", scale=1/1e6)
    ll_2324 = extract_from_sheet("DRE Gerencial 2023 a 2024", "Lucro líquido", scale=1/1000)
    ll_curr = extract_from_sheet("DRE Gerencial", "Lucro líquido", scale=1/1000)

    mf_old = extract_from_sheet("DRE_Gerencial até 2018", "Margem Financeira", scale=1/1e6)
    mf_curr = extract_from_sheet("DRE Gerencial", "Margem Financeira Bruta", scale=1/1000)

    # Carteira e Balanço
    cart_hist = extract_from_sheet("Crédito_Segmento 2014 a 2024", "Total", scale=1/1000)
    cart_curr = extract_from_sheet("Crédito Segmento", "Total", scale=1/1000)
    bp_hist = extract_from_sheet("Balanço Patrimonial até 2024", "Total do Ativo", scale=1/1000)
    bp_curr = extract_from_sheet("Balanço", "Total do Ativo", scale=1/1000)

    data = {
        "banco": "Santander Brasil",
        "ticker": "SANB11",
        "kpis": {
            "lucro_liquido_recorrente": merge_series(ll_old, ll_1922, ll_2324, ll_curr),
            "roe": merge_series(roe_2022, roe_20242, roe_curr),
            "indice_eficiencia": merge_series(ie_2022, ie_2024, ie_curr),
            "margem_financeira": merge_series(mf_old, mf_curr),
            "carteira_credito": merge_series(cart_hist, cart_curr),
            "total_ativos": merge_series(bp_hist, bp_curr),
        }
    }

    wb.close()
    return data


# ─────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────
if __name__ == "__main__":
    extractors = [
        ("itau", extract_itau),
        ("bradesco", extract_bradesco),
        ("bb", extract_bb),
        ("btg", extract_btg),
        ("santander", extract_santander),
    ]

    all_banks = []
    for name, fn in extractors:
        print(f"Extraindo {name}...")
        try:
            data = fn()
            out_path = DATA_OUT / f"{name}.json"
            with open(out_path, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            # Stats
            for kpi, series in data["kpis"].items():
                n = len([v for v in series.values() if v is not None])
                print(f"  {kpi}: {n} períodos")
            all_banks.append(data)
            print(f"  ✓ salvo em {out_path}")
        except Exception as e:
            print(f"  ✗ ERRO: {e}")
            import traceback; traceback.print_exc()

    # Arquivo consolidado
    out_all = DATA_OUT / "all_banks.json"
    with open(out_all, "w", encoding="utf-8") as f:
        json.dump(all_banks, f, ensure_ascii=False, indent=2)
    print(f"\n✓ Consolidado em {out_all}")
